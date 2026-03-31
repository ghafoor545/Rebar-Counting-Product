# backend/routes/detection_routes.py

from typing import Any, Dict, List, Optional
import os
import io
import json
from uuid import uuid4
from pathlib import Path
from datetime import datetime

import cv2
import numpy as np
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from backend.core.config import PER_PAGE
from backend.db import get_conn
from backend.utils.utils import utc_now_iso
from backend.services.detector import (
    RebarBundleDetector,
    model,
    img_to_data_uri,
    file_to_data_uri,
)
from backend.services.oak_utils import grab_oak_frame  # Updated import
from backend.oled_display import (
    oled_show_processing,
    oled_show_count,
    oled_show_message,
)

router = APIRouter(prefix="/detections", tags=["detections"])

# OAK session store
OAK_SESSION: Dict[str, Any] = {}

# Detection service instance with depth filtering enabled
detector_service = RebarBundleDetector(
    eps=100.0,
    min_bundle_size=5,
    min_samples=2,
    row_tolerance=40.0,
    use_adaptive_eps=True,
    # Enable depth filtering for OAK-D Pro
    use_depth_filter=True,
    max_detection_distance_mm=1500.0,  # 1.5 meters - adjust as needed
    min_detection_distance_mm=200.0,   # 20 cm
    debug=False,  # Set True to see depth filter logs
)

# -------------------------
# Paths for saved detections
# -------------------------
BACKEND_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BACKEND_DIR / "data"
IMAGES_DIR = DATA_DIR / "images"
THUMBS_DIR = DATA_DIR / "thumbs"

IMAGES_DIR.mkdir(parents=True, exist_ok=True)
THUMBS_DIR.mkdir(parents=True, exist_ok=True)


def _row_to_dict(row: Any) -> Dict[str, Any]:
    if row is None:
        return {}
    try:
        return dict(row)
    except Exception:
        return {"value": row}


# -------------------------
# DB helper functions
# -------------------------
def record_detection(
    user_id: int,
    processed_rgb: np.ndarray,
    count: int,
    stream_url: str,
    snapshot_url: str,
    bundle_info: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Save annotated image, thumbnail and metadata to PostgreSQL.
    Returns the detection ID (UUID string).
    """
    det_id = str(uuid4())

    # Ensure RGB -> BGR for OpenCV
    if processed_rgb.ndim == 3 and processed_rgb.shape[2] == 3:
        img_bgr = cv2.cvtColor(processed_rgb, cv2.COLOR_RGB2BGR)
    else:
        img_bgr = processed_rgb

    h, w = processed_rgb.shape[:2]
    width, height = int(w), int(h)

    # Save full image
    image_path = IMAGES_DIR / f"{det_id}.jpg"
    cv2.imwrite(
        str(image_path),
        img_bgr,
        [cv2.IMWRITE_JPEG_QUALITY, 90],
    )

    # Save thumbnail
    max_thumb_w = 320
    scale = min(1.0, max_thumb_w / float(width)) if width > 0 else 1.0
    thumb_size = (max(int(width * scale), 1), max(int(height * scale), 1))
    thumb_bgr = cv2.resize(img_bgr, thumb_size, interpolation=cv2.INTER_AREA)

    thumb_path = THUMBS_DIR / f"{det_id}_thumb.jpg"
    cv2.imwrite(
        str(thumb_path),
        thumb_bgr,
        [cv2.IMWRITE_JPEG_QUALITY, 85],
    )

    bundle_text = json.dumps(bundle_info) if bundle_info is not None else None

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO detections (
            id,
            user_id,
            timestamp,
            stream_url,
            snapshot_url,
            image_path,
            thumb_path,
            count,
            width,
            height,
            bundle_info
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            det_id,
            user_id,
            utc_now_iso(),
            stream_url,
            snapshot_url,
            str(image_path),
            str(thumb_path),
            int(count),
            width,
            height,
            bundle_text,
        ),
    )
    conn.commit()
    conn.close()

    print(f"[record_detection] user_id={user_id}, det_id={det_id}")
    return det_id


def list_detections(user_id: int, page: int, per_page: int):
    """
    Return (rows, total_count) for a user's detections, newest first.
    """
    offset = max(page - 1, 0) * per_page

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT *
        FROM detections
        WHERE user_id = %s
        ORDER BY timestamp DESC
        LIMIT %s OFFSET %s
        """,
        (user_id, per_page, offset),
    )
    rows = cur.fetchall()

    cur.execute(
        "SELECT COUNT(*) AS c FROM detections WHERE user_id = %s",
        (user_id,),
    )
    total = cur.fetchone()["c"]
    conn.close()

    print(
        f"[list_detections] user_id={user_id}, page={page}, per_page={per_page}, rows={len(rows)}, total={total}"
    )
    return rows, total


def get_detection(det_id: str, user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM detections WHERE id = %s AND user_id = %s LIMIT 1",
        (det_id, user_id),
    )
    row = cur.fetchone()
    conn.close()
    return row


def delete_detection(det_id: str, user_id: int) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM detections WHERE id = %s AND user_id = %s",
        (det_id, user_id),
    )
    deleted = cur.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


# -----------------------------
# Capture & Count from IP Webcam
# -----------------------------
@router.post("/capture/ip")
def capture_from_ip(
    user_id: int = Form(...),
    stream_url: str = Form(...),
    snapshot_url: str = Form(...),
):
    try:
        oled_show_processing()
    except Exception:
        pass

    img_bgr, err = detector_service.fetch_snapshot(snapshot_url)
    if err or img_bgr is None:
        try:
            oled_show_message("IP Error", (err or "")[:12])
        except Exception:
            pass

        return {
            "det_id": None,
            "count": 0,
            "error": err or "Failed to fetch snapshot.",
            "bundle_info": None,
            "image": None,
            "stream_url": stream_url,
            "snapshot_url": snapshot_url,
        }

    # No depth data for IP webcam
    annotated_rgb, count, derr, bundle_info = detector_service.detect_rebars(
        img_bgr, model, depth_map=None, conf=0.5, iou=0.3, max_det=10000
    )
    if derr:
        try:
            oled_show_message("Detect Error", derr[:12])
        except Exception:
            pass

        return {
            "det_id": None,
            "count": 0,
            "error": derr,
            "bundle_info": None,
            "image": None,
            "stream_url": stream_url,
            "snapshot_url": snapshot_url,
        }

    det_id = record_detection(
        user_id=user_id,
        processed_rgb=annotated_rgb,
        count=count,
        stream_url=stream_url,
        snapshot_url=snapshot_url,
        bundle_info=bundle_info,
    )

    try:
        oled_show_count(count)
    except Exception:
        pass

    image_uri = img_to_data_uri(annotated_rgb, quality=88, max_w=720)

    return {
        "det_id": det_id,
        "count": int(count),
        "error": None,
        "bundle_info": bundle_info,
        "image": image_uri,
        "stream_url": stream_url,
        "snapshot_url": snapshot_url,
    }


# ---------------------------------------------------------
# Capture & Count from OAK-D (WITH DEPTH FILTERING) ✅
# ---------------------------------------------------------
@router.post("/capture/oak")
def capture_from_oak(user_id: int = Form(...)):
    """
    Capture & Count from OAK-D Pro with depth filtering.
    """
    try:
        oled_show_processing()
    except Exception:
        pass

    # Updated: grab_oak_frame now returns (rgb_frame, depth_map, error)
    frame, depth_map, oerr = grab_oak_frame(OAK_SESSION, wait_sec=2.0)
    
    if oerr or frame is None:
        try:
            oled_show_message("OAK Error", (oerr or "")[:12])
        except Exception:
            pass

        return {
            "det_id": None,
            "count": 0,
            "error": oerr or "Failed to grab frame from OAK-D Pro.",
            "bundle_info": None,
            "image": None,
            "stream_source": "OAK-D Pro",
        }

    # Pass depth_map to enable depth filtering
    annotated_rgb, count, derr, bundle_info = detector_service.detect_rebars(
        frame, model, depth_map=depth_map, conf=0.5, iou=0.3, max_det=10000
    )
    
    if derr:
        try:
            oled_show_message("Detect Error", derr[:12])
        except Exception:
            pass

        return {
            "det_id": None,
            "count": 0,
            "error": derr,
            "bundle_info": None,
            "image": None,
            "stream_source": "OAK-D Pro",
        }

    det_id = record_detection(
        user_id=user_id,
        processed_rgb=annotated_rgb,
        count=count,
        stream_url="OAK-D Pro",
        snapshot_url="OAK-D Pro",
        bundle_info=bundle_info,
    )

    try:
        oled_show_count(count)
    except Exception:
        pass

    image_uri = img_to_data_uri(annotated_rgb, quality=88, max_w=720)

    return {
        "det_id": det_id,
        "count": int(count),
        "error": None,
        "bundle_info": bundle_info,
        "image": image_uri,
        "stream_source": "OAK-D Pro",
    }


# ----------------------------------------
# Detect from uploaded image + save record
# ----------------------------------------
@router.post("/upload")
async def detect_uploaded_image(
    user_id: int = Form(...),
    file: UploadFile = File(...),
    stream_url: Optional[str] = Form(None),
):
    try:
        oled_show_processing()
    except Exception:
        pass

    contents = await file.read()
    if not contents:
        try:
            oled_show_message("Upload Error", "No file")
        except Exception:
            pass

        return {
            "det_id": None,
            "count": 0,
            "error": "No file uploaded.",
            "bundle_info": None,
            "image": None,
            "source": "Upload",
        }

    nparr = np.frombuffer(contents, np.uint8)
    img_bgr = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img_bgr is None:
        try:
            oled_show_message("Decode Error", "Bad file")
        except Exception:
            pass

        return {
            "det_id": None,
            "count": 0,
            "error": "Could not decode the uploaded image.",
            "bundle_info": None,
            "image": None,
            "source": "Upload",
        }

    # No depth data for uploaded images
    annotated_rgb, count, derr, bundle_info = detector_service.detect_rebars(
        img_bgr, model, depth_map=None, conf=0.5, iou=0.3, max_det=10000
    )
    
    if derr:
        try:
            oled_show_message("Detect Error", derr[:12])
        except Exception:
            pass

        return {
            "det_id": None,
            "count": 0,
            "error": derr,
            "bundle_info": None,
            "image": None,
            "source": "Upload",
        }

    source_label = stream_url if stream_url else "Upload"
    filename = file.filename or "Uploaded Image"

    det_id = record_detection(
        user_id=user_id,
        processed_rgb=annotated_rgb,
        count=count,
        stream_url=source_label,
        snapshot_url=filename,
        bundle_info=bundle_info,
    )

    try:
        oled_show_count(count)
    except Exception:
        pass

    image_uri = img_to_data_uri(annotated_rgb, quality=88, max_w=720)

    return {
        "det_id": det_id,
        "count": int(count),
        "error": None,
        "bundle_info": bundle_info,
        "image": image_uri,
        "source": source_label,
    }


# -----------------------------------------
# Recent detections for gallery (simple list)
# -----------------------------------------
@router.get("/recent")
def recent_detections(
    user_id: int,
    limit: int = 48,
):
    rows, total = list_detections(user_id, page=1, per_page=limit)
    items: List[Dict[str, Any]] = []

    for r in rows:
        rd = _row_to_dict(r)
        image_uri = file_to_data_uri(rd.get("image_path", ""), max_w=720, quality=88)
        rd["image"] = image_uri
        items.append(rd)

    return {
        "items": items,
        "returned": len(items),
        "total_available": int(total),
    }


# ------------------------------
# History: list detections (table)
# ------------------------------
@router.get("")
def list_user_detections(
    user_id: int,
    page: int = 1,
    per_page: int = PER_PAGE,
):
    rows, total = list_detections(user_id, page, per_page)
    items: List[Dict[str, Any]] = []

    for r in rows:
        rd = _row_to_dict(r)
        thumb_uri = file_to_data_uri(rd.get("thumb_path", ""), max_w=90, quality=85)
        rd["thumb_uri"] = thumb_uri
        items.append(rd)

    return {
        "items": items,
        "total": int(total),
        "page": int(page),
        "per_page": int(per_page),
    }


# -----------------------------------------
# Export detections history to Excel
# -----------------------------------------
@router.get("/export")
def export_detections_excel(
    user_id: int,
    limit: int = 1000,
):
    """
    Export user's detections to an Excel file (XLSX) with thumbnail images.
    """
    rows, total = list_detections(user_id, page=1, per_page=limit)

    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"

    headers = [
        "ID",
        "Timestamp",
        "Stream",
        "Snapshot",
        "Count",
        "Bundles",
        "Isolated",
        "Image",
    ]
    ws.append(headers)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 22

    start_row = 2
    img_col_letter = "H"

    target_img_width_px = 120
    target_img_height_px = 80

    for idx, r in enumerate(rows):
        rd = _row_to_dict(r)
        row_index = start_row + idx

        display_id = idx + 1

        bi_raw = rd.get("bundle_info")
        if isinstance(bi_raw, str):
            try:
                bi = json.loads(bi_raw)
            except Exception:
                bi = {}
        else:
            bi = bi_raw or {}

        bundles = bi.get("total_bundles", 0) or 0
        in_bundles = bi.get("rebars_in_bundles", 0) or 0
        isolated = bi.get("isolated") or bi.get("total_isolated") or 0

        db_count = rd.get("count", 0) or 0
        display_count = db_count

        ts_raw = rd.get("timestamp")
        ts_display = ts_raw
        if ts_raw:
            try:
                ts_norm = ts_raw.replace("Z", "+00:00")
                dt = datetime.fromisoformat(ts_norm)
                ts_display = dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                ts_display = ts_raw

        ws.cell(row=row_index, column=1, value=display_id)
        ws.cell(row=row_index, column=2, value=ts_display)
        ws.cell(row=row_index, column=3, value=rd.get("stream_url"))
        ws.cell(row=row_index, column=4, value=rd.get("snapshot_url"))
        ws.cell(row=row_index, column=5, value=display_count)
        ws.cell(row=row_index, column=6, value=bundles)
        ws.cell(row=row_index, column=7, value=isolated)

        ws.row_dimensions[row_index].height = target_img_height_px * 0.75

        image_path = rd.get("thumb_path") or rd.get("image_path")
        if not image_path or not os.path.exists(image_path):
            continue

        try:
            img = XLImage(image_path)
        except Exception:
            continue

        img.width = target_img_width_px
        img.height = target_img_height_px
        img.anchor = f"{img_col_letter}{row_index}"
        ws.add_image(img)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="detections.xlsx"'
    }

    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )


# -----------------------------------
# Single detection detail + full image
# -----------------------------------
@router.get("/{det_id}")
def get_detection_detail(det_id: str, user_id: int):
    row = get_detection(det_id, user_id)
    if not row:
        raise HTTPException(
            status_code=404, detail="Detection not found."
        )

    rd = _row_to_dict(row)
    image_uri = file_to_data_uri(rd.get("image_path", ""), max_w=1280, quality=90)

    return {
        "detection": rd,
        "image": image_uri,
    }


# -------------------
# Delete detection row
# -------------------
@router.delete("/{det_id}")
def delete_detection_entry(det_id: str, user_id: int):
    ok = delete_detection(det_id, user_id)
    if not ok:
        raise HTTPException(
            status_code=404, detail="Detection not found."
        )
    return {"success": True}
